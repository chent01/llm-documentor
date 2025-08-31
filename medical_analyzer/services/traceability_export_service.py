"""
Traceability matrix export service with multiple format support.

This service provides comprehensive export capabilities for traceability matrices
including CSV, Excel, and PDF formats with proper formatting and gap indicators.
"""

import logging
import csv
import json
import os
from typing import List, Dict, Any, Optional
from datetime import datetime
from io import StringIO
import tempfile

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from .traceability_models import TraceabilityMatrix, TraceabilityTableRow, TraceabilityGap


logger = logging.getLogger(__name__)


class TraceabilityExportService:
    """Service for exporting traceability matrices in multiple formats."""
    
    def __init__(self):
        """Initialize the export service."""
        self.logger = logging.getLogger(__name__)
        
    def export_csv(
        self, 
        table_rows: List[TraceabilityTableRow], 
        gaps: List[TraceabilityGap],
        filename: str,
        include_gaps: bool = True
    ) -> bool:
        """
        Export traceability matrix to CSV format.
        
        Args:
            table_rows: List of traceability table rows
            gaps: List of detected gaps
            filename: Output filename
            include_gaps: Whether to include gap indicators
            
        Returns:
            True if export successful, False otherwise
        """
        try:
            self.logger.info(f"Exporting traceability matrix to CSV: {filename}")
            
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header
                headers = [
                    "Code Reference",
                    "File Path",
                    "Function Name",
                    "Feature ID",
                    "Feature Description",
                    "User Requirement ID",
                    "User Requirement Text",
                    "Software Requirement ID",
                    "Software Requirement Text",
                    "Risk ID",
                    "Risk Hazard",
                    "Confidence"
                ]
                
                if include_gaps:
                    headers.extend(["Gap Indicators", "Gap Severity", "Gap Recommendations"])
                
                writer.writerow(headers)
                
                # Create gap lookup for efficient access
                gap_lookup = self._create_gap_lookup(gaps) if include_gaps else {}
                
                # Write data rows
                for row in table_rows:
                    csv_row = [
                        row.code_reference,
                        row.file_path,
                        row.function_name,
                        row.feature_id,
                        row.feature_description,
                        row.user_requirement_id,
                        row.user_requirement_text,
                        row.software_requirement_id,
                        row.software_requirement_text,
                        row.risk_id,
                        row.risk_hazard,
                        f"{row.confidence:.2f}"
                    ]
                    
                    if include_gaps:
                        gap_info = self._get_row_gap_info(row, gap_lookup)
                        csv_row.extend([
                            gap_info["indicators"],
                            gap_info["severity"],
                            gap_info["recommendations"]
                        ])
                    
                    writer.writerow(csv_row)
                
                # Write gap summary if requested
                if include_gaps and gaps:
                    writer.writerow([])  # Empty row
                    writer.writerow(["GAP ANALYSIS SUMMARY"])
                    writer.writerow(["Gap Type", "Count", "Severity", "Description"])
                    
                    gap_summary = self._summarize_gaps(gaps)
                    for gap_type, info in gap_summary.items():
                        writer.writerow([
                            gap_type.replace('_', ' ').title(),
                            info["count"],
                            info["max_severity"],
                            info["description"]
                        ])
            
            self.logger.info(f"CSV export completed successfully: {filename}")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to export CSV: {e}")
            return False
            
    def export_excel(
        self,
        table_rows: List[TraceabilityTableRow],
        gaps: List[TraceabilityGap],
        filename: str,
        include_formatting: bool = True
    ) -> bool:
        """
        Export traceability matrix to Excel format with conditional formatting.
        
        Args:
            table_rows: List of traceability table rows
            gaps: List of detected gaps
            filename: Output filename
            include_formatting: Whether to apply conditional formatting
            
        Returns:
            True if export successful, False otherwise
        """
        if not EXCEL_AVAILABLE:
            self.logger.error("Excel export not available - openpyxl not installed")
            return False
            
        try:
            self.logger.info(f"Exporting traceability matrix to Excel: {filename}")
            
            # Create workbook and worksheets
            wb = openpyxl.Workbook()
            
            # Main matrix sheet
            ws_matrix = wb.active
            ws_matrix.title = "Traceability Matrix"
            
            # Gap analysis sheet
            ws_gaps = wb.create_sheet("Gap Analysis")
            
            # Export main matrix
            self._export_excel_matrix(ws_matrix, table_rows, gaps, include_formatting)
            
            # Export gap analysis
            self._export_excel_gaps(ws_gaps, gaps, include_formatting)
            
            # Save workbook
            wb.save(filename)
            
            self.logger.info(f"Excel export completed successfully: {filename}")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to export Excel: {e}")
            return False
            
    def _export_excel_matrix(
        self,
        worksheet,
        table_rows: List[TraceabilityTableRow],
        gaps: List[TraceabilityGap],
        include_formatting: bool
    ):
        """Export matrix data to Excel worksheet."""
        # Headers
        headers = [
            "Code Reference", "File Path", "Function Name", "Feature ID",
            "Feature Description", "User Requirement ID", "User Requirement Text",
            "Software Requirement ID", "Software Requirement Text", "Risk ID",
            "Risk Hazard", "Confidence", "Gap Status"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            if include_formatting:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Create gap lookup
        gap_lookup = self._create_gap_lookup(gaps)
        
        # Write data rows
        for row_idx, row in enumerate(table_rows, 2):
            data = [
                row.code_reference,
                row.file_path,
                row.function_name,
                row.feature_id,
                row.feature_description,
                row.user_requirement_id,
                row.user_requirement_text,
                row.software_requirement_id,
                row.software_requirement_text,
                row.risk_id,
                row.risk_hazard,
                row.confidence,
                self._get_row_gap_status(row, gap_lookup)
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                
                if include_formatting:
                    # Apply gap highlighting
                    gap_severity = self._get_row_gap_severity(row, gap_lookup)
                    if gap_severity:
                        if gap_severity == "high":
                            cell.fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
                        elif gap_severity == "medium":
                            cell.fill = PatternFill(start_color="FFFFD6", end_color="FFFFD6", fill_type="solid")
                        elif gap_severity == "low":
                            cell.fill = PatternFill(start_color="D6FFD6", end_color="D6FFD6", fill_type="solid")
                    
                    # Confidence color coding
                    if col_idx == 12:  # Confidence column
                        if row.confidence >= 0.8:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif row.confidence >= 0.5:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Auto-adjust column widths
        if include_formatting:
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
    def _export_excel_gaps(self, worksheet, gaps: List[TraceabilityGap], include_formatting: bool):
        """Export gap analysis to Excel worksheet."""
        # Headers
        headers = ["Gap Type", "Severity", "Source Type", "Source ID", "Target Type", "Target ID", "Description", "Recommendation"]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            if include_formatting:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write gap data
        for row_idx, gap in enumerate(gaps, 2):
            data = [
                gap.gap_type.replace('_', ' ').title(),
                gap.severity.title(),
                gap.source_type,
                gap.source_id,
                gap.target_type or "",
                gap.target_id or "",
                gap.description,
                gap.recommendation
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                
                if include_formatting and col_idx == 2:  # Severity column
                    if gap.severity == "high":
                        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif gap.severity == "medium":
                        cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                    elif gap.severity == "low":
                        cell.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
        
        # Auto-adjust column widths
        if include_formatting:
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
    def export_pdf(
        self,
        table_rows: List[TraceabilityTableRow],
        gaps: List[TraceabilityGap],
        filename: str,
        include_summary: bool = True
    ) -> bool:
        """
        Export traceability matrix to PDF format for regulatory submissions.
        
        Args:
            table_rows: List of traceability table rows
            gaps: List of detected gaps
            filename: Output filename
            include_summary: Whether to include executive summary
            
        Returns:
            True if export successful, False otherwise
        """
        if not PDF_AVAILABLE:
            self.logger.error("PDF export not available - reportlab not installed")
            return False
            
        try:
            self.logger.info(f"Exporting traceability matrix to PDF: {filename}")
            
            # Create PDF document
            doc = SimpleDocTemplate(filename, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            story.append(Paragraph("Traceability Matrix Report", title_style))
            story.append(Spacer(1, 12))
            
            # Metadata
            metadata_style = styles['Normal']
            story.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", metadata_style))
            story.append(Paragraph(f"<b>Total Rows:</b> {len(table_rows)}", metadata_style))
            story.append(Paragraph(f"<b>Total Gaps:</b> {len(gaps)}", metadata_style))
            story.append(Spacer(1, 20))
            
            # Executive summary
            if include_summary:
                self._add_pdf_summary(story, table_rows, gaps, styles)
                story.append(PageBreak())
            
            # Main traceability matrix
            self._add_pdf_matrix(story, table_rows, gaps, styles)
            
            # Gap analysis
            if gaps:
                story.append(PageBreak())
                self._add_pdf_gaps(story, gaps, styles)
            
            # Build PDF
            doc.build(story)
            
            self.logger.info(f"PDF export completed successfully: {filename}")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to export PDF: {e}")
            return False
            
    def _add_pdf_summary(self, story, table_rows: List[TraceabilityTableRow], gaps: List[TraceabilityGap], styles):
        """Add executive summary to PDF."""
        story.append(Paragraph("Executive Summary", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Calculate metrics
        complete_chains = sum(1 for row in table_rows 
                            if all([row.user_requirement_id, row.software_requirement_id, row.risk_id]))
        coverage_pct = (complete_chains / len(table_rows) * 100) if table_rows else 0
        
        avg_confidence = sum(row.confidence for row in table_rows) / len(table_rows) if table_rows else 0
        
        high_gaps = sum(1 for gap in gaps if gap.severity == "high")
        medium_gaps = sum(1 for gap in gaps if gap.severity == "medium")
        low_gaps = sum(1 for gap in gaps if gap.severity == "low")
        
        # Summary content
        summary_data = [
            ["Metric", "Value", "Status"],
            ["Total Traceability Links", str(len(table_rows)), ""],
            ["Complete Chains", f"{complete_chains} ({coverage_pct:.1f}%)", "✓" if coverage_pct >= 80 else "⚠"],
            ["Average Confidence", f"{avg_confidence:.2f}", "✓" if avg_confidence >= 0.7 else "⚠"],
            ["High Severity Gaps", str(high_gaps), "✓" if high_gaps == 0 else "⚠"],
            ["Medium Severity Gaps", str(medium_gaps), ""],
            ["Low Severity Gaps", str(low_gaps), ""]
        ]
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 1.5*inch, 0.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Key findings
        story.append(Paragraph("Key Findings", styles['Heading3']))
        findings = []
        
        if coverage_pct < 50:
            findings.append("• Low traceability coverage indicates significant gaps in requirement implementation")
        elif coverage_pct < 80:
            findings.append("• Moderate traceability coverage with room for improvement")
        else:
            findings.append("• Good traceability coverage indicating well-structured implementation")
            
        if high_gaps > 0:
            findings.append(f"• {high_gaps} high-severity gaps require immediate attention")
            
        if avg_confidence < 0.5:
            findings.append("• Low average confidence suggests need for stronger traceability evidence")
            
        for finding in findings:
            story.append(Paragraph(finding, styles['Normal']))
            
    def _add_pdf_matrix(self, story, table_rows: List[TraceabilityTableRow], gaps: List[TraceabilityGap], styles):
        """Add traceability matrix to PDF."""
        story.append(Paragraph("Traceability Matrix", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Create gap lookup
        gap_lookup = self._create_gap_lookup(gaps)
        
        # Prepare table data (limit columns for PDF width)
        matrix_data = [["Code Ref", "File", "Feature", "User Req", "SW Req", "Risk", "Conf", "Gaps"]]
        
        for row in table_rows[:50]:  # Limit rows for PDF
            gap_status = self._get_row_gap_status(row, gap_lookup)
            matrix_data.append([
                row.code_reference[:15] + "..." if len(row.code_reference) > 15 else row.code_reference,
                os.path.basename(row.file_path),
                row.feature_id,
                row.user_requirement_id,
                row.software_requirement_id,
                row.risk_id,
                f"{row.confidence:.2f}",
                gap_status
            ])
        
        if len(table_rows) > 50:
            matrix_data.append(["...", "...", "...", "...", "...", "...", "...", f"({len(table_rows) - 50} more rows)"])
        
        # Create table
        matrix_table = Table(matrix_data, colWidths=[0.8*inch, 0.8*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.4*inch, 0.6*inch])
        matrix_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        ]))
        
        story.append(matrix_table)
        
    def _add_pdf_gaps(self, story, gaps: List[TraceabilityGap], styles):
        """Add gap analysis to PDF."""
        story.append(Paragraph("Gap Analysis", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Group gaps by severity
        high_gaps = [g for g in gaps if g.severity == "high"]
        medium_gaps = [g for g in gaps if g.severity == "medium"]
        low_gaps = [g for g in gaps if g.severity == "low"]
        
        for severity, severity_gaps in [("High", high_gaps), ("Medium", medium_gaps), ("Low", low_gaps)]:
            if not severity_gaps:
                continue
                
            story.append(Paragraph(f"{severity} Severity Gaps ({len(severity_gaps)})", styles['Heading3']))
            
            for i, gap in enumerate(severity_gaps[:10], 1):  # Limit to 10 per severity
                gap_text = f"{i}. <b>{gap.gap_type.replace('_', ' ').title()}:</b> {gap.description}"
                story.append(Paragraph(gap_text, styles['Normal']))
                story.append(Paragraph(f"   <i>Recommendation:</i> {gap.recommendation}", styles['Normal']))
                story.append(Spacer(1, 6))
                
            if len(severity_gaps) > 10:
                story.append(Paragraph(f"... and {len(severity_gaps) - 10} more {severity.lower()} severity gaps", styles['Normal']))
                
            story.append(Spacer(1, 12))
            
    def export_gap_report(self, gaps: List[TraceabilityGap], filename: str) -> bool:
        """
        Export detailed gap analysis report to text file.
        
        Args:
            gaps: List of detected gaps
            filename: Output filename
            
        Returns:
            True if export successful, False otherwise
        """
        try:
            self.logger.info(f"Exporting gap report to: {filename}")
            
            with open(filename, 'w', encoding='utf-8') as f:
                f.write("TRACEABILITY GAP ANALYSIS REPORT\n")
                f.write("=" * 50 + "\n\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Total Gaps: {len(gaps)}\n\n")
                
                # Summary by severity
                severity_counts = {}
                for gap in gaps:
                    severity_counts[gap.severity] = severity_counts.get(gap.severity, 0) + 1
                
                f.write("SUMMARY BY SEVERITY\n")
                f.write("-" * 20 + "\n")
                for severity in ["high", "medium", "low"]:
                    count = severity_counts.get(severity, 0)
                    f.write(f"{severity.title()}: {count}\n")
                f.write("\n")
                
                # Summary by type
                type_counts = {}
                for gap in gaps:
                    type_counts[gap.gap_type] = type_counts.get(gap.gap_type, 0) + 1
                
                f.write("SUMMARY BY TYPE\n")
                f.write("-" * 15 + "\n")
                for gap_type, count in type_counts.items():
                    f.write(f"{gap_type.replace('_', ' ').title()}: {count}\n")
                f.write("\n")
                
                # Detailed gaps
                for severity in ["high", "medium", "low"]:
                    severity_gaps = [g for g in gaps if g.severity == severity]
                    if not severity_gaps:
                        continue
                        
                    f.write(f"{severity.upper()} SEVERITY GAPS\n")
                    f.write("-" * 25 + "\n")
                    
                    for i, gap in enumerate(severity_gaps, 1):
                        f.write(f"{i}. {gap.description}\n")
                        f.write(f"   Type: {gap.gap_type}\n")
                        f.write(f"   Source: {gap.source_type} '{gap.source_id}'\n")
                        if gap.target_type and gap.target_id:
                            f.write(f"   Target: {gap.target_type} '{gap.target_id}'\n")
                        f.write(f"   Recommendation: {gap.recommendation}\n\n")
                    
                    f.write("\n")
            
            self.logger.info(f"Gap report export completed successfully: {filename}")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to export gap report: {e}")
            return False
            
    def _create_gap_lookup(self, gaps: List[TraceabilityGap]) -> Dict[str, List[TraceabilityGap]]:
        """Create lookup dictionary for gaps by source ID."""
        lookup = {}
        for gap in gaps:
            key = f"{gap.source_type}:{gap.source_id}"
            if key not in lookup:
                lookup[key] = []
            lookup[key].append(gap)
        return lookup
        
    def _get_row_gap_info(self, row: TraceabilityTableRow, gap_lookup: Dict[str, List[TraceabilityGap]]) -> Dict[str, str]:
        """Get gap information for a table row."""
        # Check for gaps related to this row
        row_gaps = []
        
        # Check code reference gaps
        code_key = f"code:{row.code_reference}"
        row_gaps.extend(gap_lookup.get(code_key, []))
        
        # Check feature gaps
        feature_key = f"feature:{row.feature_id}"
        row_gaps.extend(gap_lookup.get(feature_key, []))
        
        # Check requirement gaps
        if row.user_requirement_id:
            ur_key = f"requirement:{row.user_requirement_id}"
            row_gaps.extend(gap_lookup.get(ur_key, []))
            
        if row.software_requirement_id:
            sr_key = f"requirement:{row.software_requirement_id}"
            row_gaps.extend(gap_lookup.get(sr_key, []))
        
        # Check for missing elements (implicit gaps)
        implicit_gaps = []
        if not row.user_requirement_id:
            implicit_gaps.append("Missing UR")
        if not row.software_requirement_id:
            implicit_gaps.append("Missing SR")
        if not row.risk_id:
            implicit_gaps.append("Missing Risk")
        if row.confidence < 0.5:
            implicit_gaps.append("Low Confidence")
        
        # Combine gap information
        gap_types = [gap.gap_type for gap in row_gaps] + implicit_gaps
        severities = [gap.severity for gap in row_gaps]
        recommendations = [gap.recommendation for gap in row_gaps]
        
        max_severity = "low"
        if "high" in severities or implicit_gaps:
            max_severity = "high"
        elif "medium" in severities:
            max_severity = "medium"
        
        return {
            "indicators": "; ".join(gap_types) if gap_types else "None",
            "severity": max_severity if gap_types else "none",
            "recommendations": "; ".join(recommendations[:2]) if recommendations else "None"  # Limit length
        }
        
    def _get_row_gap_status(self, row: TraceabilityTableRow, gap_lookup: Dict[str, List[TraceabilityGap]]) -> str:
        """Get concise gap status for a row."""
        gap_info = self._get_row_gap_info(row, gap_lookup)
        if gap_info["severity"] == "none":
            return "OK"
        elif gap_info["severity"] == "high":
            return "HIGH"
        elif gap_info["severity"] == "medium":
            return "MED"
        else:
            return "LOW"
            
    def _get_row_gap_severity(self, row: TraceabilityTableRow, gap_lookup: Dict[str, List[TraceabilityGap]]) -> Optional[str]:
        """Get the highest gap severity for a row."""
        gap_info = self._get_row_gap_info(row, gap_lookup)
        return gap_info["severity"] if gap_info["severity"] != "none" else None
        
    def _summarize_gaps(self, gaps: List[TraceabilityGap]) -> Dict[str, Dict[str, Any]]:
        """Summarize gaps by type."""
        summary = {}
        
        for gap in gaps:
            if gap.gap_type not in summary:
                summary[gap.gap_type] = {
                    "count": 0,
                    "severities": [],
                    "description": ""
                }
            
            summary[gap.gap_type]["count"] += 1
            summary[gap.gap_type]["severities"].append(gap.severity)
        
        # Add descriptions and max severity
        for gap_type, info in summary.items():
            severities = info["severities"]
            if "high" in severities:
                info["max_severity"] = "High"
            elif "medium" in severities:
                info["max_severity"] = "Medium"
            else:
                info["max_severity"] = "Low"
                
            # Add description based on gap type
            descriptions = {
                "orphaned_code": "Code without feature links",
                "orphaned_feature": "Features without requirement links",
                "orphaned_requirement": "Requirements without proper links",
                "orphaned_risk": "Risks without requirement links",
                "missing_link": "Expected links not found",
                "weak_link": "Links with low confidence",
                "broken_chain": "Incomplete traceability chains",
                "duplicate_link": "Duplicate traceability links"
            }
            info["description"] = descriptions.get(gap_type, "Unknown gap type")
        
        return summary